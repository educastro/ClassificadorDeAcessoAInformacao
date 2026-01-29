# Bibliotecas utilizadas e importadas
import os
import re
import json
import time
import argparse
from typing import Optional, Tuple, Dict, Any
import openpyxl
from tenacity import retry, wait_exponential, stop_after_attempt
from openai import OpenAI

# Expressão regular utilizada para identificação de endereços de e-mail
EMAIL_RE = re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b")
# Expressão regular utilizada para identificação de CPFs
CPF_RE = re.compile(r"\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b")
# Expressão regular utilizada para identificação de telefones
PHONE_RE = re.compile(r"(?:\+?55\s*)?(?:\(?\d{2}\)?\s*)?(?:9?\d{4})-?\d{4}\b")
# Expressão regular utilizada para identificação de RGs
RG_RE = re.compile(r"\b\d{1,2}\.?\d{3}\.?\d{2,3}-?[0-9Xx]\b") 


# Função que verifica se há dados pessoais com base nas expressões regulares
def has_personal_data(text: str) -> bool:
    if not text:
        return False
    t = text.strip()
    return bool(
        EMAIL_RE.search(t) or CPF_RE.search(t) or PHONE_RE.search(t) or RG_RE.search(t)
    )

# Prompt para verificação se há dados pessoais com base na inteligência artificial
SYSTEM_INSTRUCTIONS = (
    "Você é um classificador de texto para pedidos de acesso à informação. "
    "Seu objetivo é identificar se o texto contém DADOS PESSOAIS que permitam a "
    "identificação direta ou indireta de pessoa natural, tais como: nome de pessoa, "
    "CPF, RG, telefone e endereço de e-mail. Os casos que conterem algum tipo de dado "
    "pessoal deverão ser classificados como \"contendo dados pessoais\", e aqueles que não possuem "
    "dado pessoal enquanto \"PUBLICO\". Também classifique como \"contendo dados pessoais\" todos os pedidos "
    "cujo o resultado envolva dados pessoais que possam identificar uma pessoa, como endereço, "
    "cadastros, laudos acerca de pessoas ou qualquer informação relativa à pessoas. Para fins "
    "de referência, considere a Lei de Acesso a Informação, a Lei n˚ 12.527, de 18 de novembro "
    "de 2011, em especial o que diz o seu artigo 31. Casos que se tratam de dúvidas simples cuja "
    "resposta não necessite de informações pessoais devem ser classificados como \"PUBLICO\". "
    "Solicitações de processos que não sejam explicitamente casos pessoais ou que tratem de "
    "dados abertos também devem ser classificados como \"PUBLICO\". Solicitações que tratem de "
    "questões e demandas pessoais, como vaga de filho na creche, atendimento na rede pública de "
    "saúde ou atendimentos do setor público de forma geral como troca de titularidade em contas "
    "devem ser classificados como \"contendo dados pessoais\". Responda APENAS em JSON válido, "
    "sem texto extra."
)

# User template para verificação se há dados pessoais com base na inteligência artificial
USER_TEMPLATE = """Analise o texto do pedido abaixo e responda em JSON com as chaves:
- "contains_personal_data": true/false
- "personal_data_types": lista curta (ex: ["nome","cpf","email","telefone","rg"]) ou []
- "decision": "contendo dados pessoais" se contains_personal_data=true, caso contrário "PUBLICO"

Texto do pedido:
\"\"\"{text}\"\"\"
"""

# Credencial para utilização da inteligênci artificial
def get_client() -> OpenAI:
    api_key = "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"
    return OpenAI(api_key=api_key)

# Função que classifica se há dados pessoais com base na inteligência artificial
@retry(wait=wait_exponential(multiplier=1, min=2, max=20), stop=stop_after_attempt(4))
def chatgpt_classify(
    client: OpenAI,
    model: str,
    text: str,
    timeout_sleep: float = 0.15,
) -> Dict[str, Any]:
    """
    Chama ChatGPT e retorna dict com:
      contains_personal_data (bool)
      personal_data_types (list)
      decision ("PUBLICO"|"contendo dados pessoais")
    """
    user_msg = USER_TEMPLATE.format(text=text)

    # Força saída estruturada via JSON Schema (Responses API)
    resp = client.responses.create(
        model=model,
        input=[
            {"role": "system", "content": SYSTEM_INSTRUCTIONS},
            {"role": "user", "content": user_msg},
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "pii_classification",
                "schema": {
                    "type": "object",
                    "properties": {
                        "contains_personal_data": {"type": "boolean"},
                        "personal_data_types": {
                            "type": "array",
                            "items": {"type": "string"},
                        },
                        "decision": {"type": "string", "enum": ["PUBLICO", "contendo dados pessoais"]},
                    },
                    "required": ["contains_personal_data", "personal_data_types", "decision"],
                    "additionalProperties": False,
                },
                "strict": True,
            }
        },
        temperature=0.0,
    )

    # A saída vem como texto JSON já validado pelo schema
    raw = resp.output_text.strip()
    data = json.loads(raw)

    time.sleep(timeout_sleep)
    return data

# Função que retorn se o texto contém não ou dado público. 
# Aqui, primeiro faz-se a verificação usando expressões regulares por ser computacionalmente mais barato
# Em seguida, caso não haja compatibilidade com alguma das expressões regulares, faz-se a verificação com IA
def decide_label(
    client: OpenAI,
    text: str,
    model: str,
) -> Tuple[str, Optional[Dict[str, Any]], str]:
    """
    Retorna:
      - label final: "PÚBLICO" ou "contendo dados pessoais"
      - json do chatgpt (ou None)
      - origem: "regex" ou "chatgpt"
    """

    # Verificação com expressões regulares
    if has_personal_data(text):
        return "contendo dados pessoais", None, "regex"

    # Verificação com IA
    data = chatgpt_classify(client=client, model=model, text=text)
    decision = (data.get("decision") or "").strip().upper()

    if decision == "contendo dados pessoais":
        return "contendo dados pessoais", data, "chatgpt"
    return "PÚBLICO", data, "chatgpt"

# Função principal
def main():
    # Parser contendo a configuração padrão do modelo
    parser = argparse.ArgumentParser(
        description="Classifica pedidos e-SIC (Público vs contendo dados pessoais) usando API do ChatGPT e escreve na coluna C."
    )
    parser.add_argument("--input", default="dados/AMOSTRA_e-SIC.xlsx", help="Caminho do XLSX de entrada")
    parser.add_argument("--output", default="dados/AMOSTRA_e-SIC_classificada.xlsx", help="Caminho do XLSX de saída")
    parser.add_argument("--sheet", default="Amostra - SIC", help="Nome da aba")
    parser.add_argument("--model", default="gpt-4.1-mini", help="Modelo (ex: gpt-4.1-mini, gpt-4.1)")
    parser.add_argument("--col_text", default=2, type=int, help="Coluna com o texto (1=A, 2=B, ...)")
    parser.add_argument("--col_out", default=3, type=int, help="Coluna de saída (1=A, 2=B, 3=C, ...)")
    parser.add_argument("--sleep_every", default=0, type=int, help="Pausa extra (s) a cada N linhas (0 desliga)")
    args = parser.parse_args()

    # Autenticação com a IA do ChatGPT
    client = get_client()

    # Carga do arquivo excel
    wb = openpyxl.load_workbook(args.input)

    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"ERRO: aba '{args.sheet}' não encontrada. Abas disponíveis: {wb.sheetnames}")
    ws = wb[args.sheet]

    # Cabeçalho na coluna de saída
    header_cell = ws.cell(row=1, column=args.col_out)
    if not header_cell.value:
        header_cell.value = "Classificação"

    max_row = ws.max_row

    # Contagem para saber quantas verificações foram feitas em regex e quantas foram feitas com IA
    regexCalls = 0
    chatGptCalls = 0
    totalCalls = 0

    public_responses = 0
    nonpublic_responses = 0

    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=args.col_text)
        text = cell.value

        # Se não há texto, nada é feito
        if not text or not str(text).strip():
            ws.cell(row=r, column=args.col_out).value = ""
            continue
        
        # Chamada da função que verifica se há ou não dados pessoais
        label, data, origin = decide_label(
            client=client,
            text=str(text),
            model=args.model,
        )

        # Contagem para saber quantas verificações foram feitas em regex e quantas foram feitas com IA
        if origin == "regex":
            regexCalls += 1
            totalCalls += 1
        elif origin == "chatgpt":
            chatGptCalls += 1
            totalCalls += 1
        else:
            totalCalls += 1
        
        if label == "PÚBLICO":
            public_responses += 1
        elif label == "NÃO PÚBLICO":
            nonpublic_responses += 1

        # Escreve o resultado no arquivo excel de saída
        ws.cell(row=r, column=args.col_out).value = label

        # Printa na tela a execução do programa, apenas para conhecimento e noção de que está rodando
        print(f"[{r}/{max_row-1}] OK (origem={origin})")

        # Pausa extra opcional para respeitar rate limit, caso exista
        if args.sleep_every and (r % args.sleep_every == 0):
            time.sleep(1)

    # Salva o arquivo excel de saída
    wb.save(args.output)

    # Imprime relatório final na tela
    print(f"Concluído.")
    print(f"Arquivo de saída: {args.output}")
    print(f"Quantidade de classificações públicas: {public_responses}")
    print(f"Quantidade de classificação que contém dados pessoais: {nonpublic_responses}")
    print(f"Quantidade de verificações por meio de regex: {regexCalls}")
    print(f"Quantidade de verificações por meio de IA: {chatGptCalls}")
    print(f"Quantidade de verificações totais: {totalCalls}")

if __name__ == "__main__":
    main()
